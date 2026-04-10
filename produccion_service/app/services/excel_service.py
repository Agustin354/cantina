import os
import logging
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from repositories.ficha_repo import (
    todos_los_movimientos, movimientos_del_periodo,
    total_fichas_emitidas, fichas_emitidas_periodo,
    total_fichas_usadas, fichas_usadas_periodo,
    total_ingresado_guaranies, ingresado_periodo,
)
from repositories.venta_repo import (
    ventas_por_producto, ventas_por_producto_periodo,
    total_ventas_guaranies, ventas_guaranies_periodo,
)
from config import REPORTS_DIR_VENTAS, REPORTS_DIR_PRODUCCION, VALOR_FICHA

logger = logging.getLogger(__name__)

COLOR_HEADER  = "366092"
COLOR_TOTAL   = "E8F5E9"
COLOR_WHITE   = "FFFFFF"
COLOR_MANUAL  = "455A64"   # gris azulado
COLOR_SEMANAL = "1565C0"   # azul
COLOR_MENSUAL = "6A1B9A"   # violeta

MESES_ES = [
    "Enero","Febrero","Marzo","Abril","Mayo","Junio",
    "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre",
]


# ── Helpers de estilo ──────────────────────────────────────────────────────────

def _header_fill(color=COLOR_HEADER):
    return PatternFill("solid", fgColor=color)

def _total_fill():
    return PatternFill("solid", fgColor=COLOR_TOTAL)

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _estilo_header(cell, color=COLOR_HEADER):
    cell.fill      = _header_fill(color)
    cell.font      = Font(color=COLOR_WHITE, bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin_border()

def _estilo_total(cell):
    cell.fill   = _total_fill()
    cell.font   = Font(bold=True)
    cell.border = _thin_border()

def _estilo_celda(cell):
    cell.border    = _thin_border()
    cell.alignment = Alignment(vertical="center")

def _fmt_moneda(cell):
    cell.number_format = "#,##0"

def _autofit(ws):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


# ── Nombre de archivo y etiqueta de período ────────────────────────────────────

def _info_reporte(tipo: str, fecha_inicio: str, fecha_fin: str):
    """Devuelve (nombre_archivo, titulo_excel, color_cabecera)."""
    now = datetime.now()
    if tipo == "semanal":
        # fecha_inicio es la fecha del lunes de la semana reportada
        dt     = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        semana = dt.isocalendar()[1]
        año    = dt.isocalendar()[0]
        fi_fmt = datetime.strptime(fecha_inicio, "%Y-%m-%d").strftime("%d/%m")
        ff_fmt = datetime.strptime(fecha_fin,    "%Y-%m-%d").strftime("%d/%m")
        nombre = f"Semanal_Sem{semana:02d}_{año}.xlsx"
        titulo = f"REPORTE SEMANAL — Semana {semana:02d}/{año}  |  Lun {fi_fmt} – Vie {ff_fmt}  |  07:00–18:00"
        color  = COLOR_SEMANAL
    elif tipo == "mensual":
        dt     = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        mes    = MESES_ES[dt.month - 1]
        nombre = f"Mensual_{mes}_{dt.year}.xlsx"
        titulo = f"REPORTE MENSUAL — {mes} {dt.year}"
        color  = COLOR_MENSUAL
    else:
        nombre = f"Manual_{now.strftime('%d-%m-%Y_%Hh%M')}.xlsx"
        titulo = f"REPORTE MANUAL — {now.strftime('%d/%m/%Y %H:%M')}"
        color  = COLOR_MANUAL
    return nombre, titulo, color


# ── Hoja 1: Resumen Financiero ─────────────────────────────────────────────────

def _hoja_resumen(wb, titulo: str, color: str,
                  fecha_inicio: str = None, fecha_fin: str = None):
    ws = wb.create_sheet("RESUMEN FINANCIERO")
    ws.sheet_view.showGridLines = False

    # Datos: filtrados por período si aplica, de lo contrario todo el historial
    if fecha_inicio and fecha_fin:
        emitidas   = fichas_emitidas_periodo(fecha_inicio, fecha_fin)
        usadas     = fichas_usadas_periodo(fecha_inicio, fecha_fin)
        ingresado  = ingresado_periodo(fecha_inicio, fecha_fin)
        gastado    = ventas_guaranies_periodo(fecha_inicio, fecha_fin)
    else:
        emitidas   = total_fichas_emitidas()
        usadas     = total_fichas_usadas()
        ingresado  = total_ingresado_guaranies()
        gastado    = total_ventas_guaranies()

    saldo_fic   = emitidas - usadas
    saldo_gs    = saldo_fic * VALOR_FICHA
    diferencia  = ingresado - gastado
    consistente = abs(diferencia) < 1

    # Título principal con color según tipo de reporte
    ws["A1"] = f"CANTINA — {titulo}"
    ws["A1"].fill      = PatternFill("solid", fgColor="1A1A2E")
    ws["A1"].font      = Font(color=COLOR_WHITE, bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 20

    filas = [
        ("CONCEPTO", "VALOR", "UNIDAD"),
        None,
        ("Total Ingresado (dinero en caja)", ingresado,  "Gs."),
        ("Total Gastado (fichas usadas)",    gastado,    "Gs."),
        ("Diferencia en caja",               diferencia, "Gs."),
        None,
        ("Fichas Emitidas",  emitidas,  "fichas"),
        ("Fichas Usadas",    usadas,    "fichas"),
        ("Saldo de Fichas",  saldo_fic, "fichas"),
        ("Saldo en Guaraníes", saldo_gs, "Gs."),
        None,
        ("Estado", "🟢 CONSISTENTE" if consistente else "🔴 HAY DIFERENCIA", ""),
    ]

    for i, fila in enumerate(filas, start=4):
        ws.row_dimensions[i].height = 22
        if fila is None:
            continue
        a, b, c = fila
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
        ws.cell(i, 3, c)
        if i == 4:
            for col in range(1, 4):
                _estilo_header(ws.cell(i, col), color)
        elif fila[0] in ("Diferencia en caja", "Saldo en Guaraníes", "Estado"):
            for col in range(1, 4):
                _estilo_total(ws.cell(i, col))
            if fila[0] in ("Diferencia en caja", "Saldo en Guaraníes"):
                _fmt_moneda(ws.cell(i, 2))
        else:
            for col in range(1, 4):
                _estilo_celda(ws.cell(i, col))
            if fila[2] == "Gs.":
                _fmt_moneda(ws.cell(i, 2))

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12


# ── Hoja 2: Movimientos de Fichas ──────────────────────────────────────────────

def _hoja_movimientos(wb, color: str,
                      fecha_inicio: str = None, fecha_fin: str = None):
    ws = wb.create_sheet("MOVIMIENTOS DE FICHAS")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24

    headers = ["Fecha", "Tipo", "Monto (Gs.)", "Cantidad Fichas", "Observación"]
    for col, h in enumerate(headers, 1):
        _estilo_header(ws.cell(1, col, h), color)

    if fecha_inicio and fecha_fin:
        movimientos = movimientos_del_periodo(fecha_inicio, fecha_fin)
    else:
        movimientos = todos_los_movimientos()

    for fila, m in enumerate(movimientos, start=2):
        ws.cell(fila, 1, m.get("fecha", ""))
        ws.cell(fila, 2, m.get("tipo", "").capitalize())
        ws.cell(fila, 3, m.get("monto_guaranies", 0))
        ws.cell(fila, 4, m.get("cantidad_fichas", 0))
        ws.cell(fila, 5, m.get("observacion", ""))
        for col in range(1, 6):
            _estilo_celda(ws.cell(fila, col))
        _fmt_moneda(ws.cell(fila, 3))

    if movimientos:
        tf = len(movimientos) + 2
        ws.cell(tf, 1, "TOTALES").font = Font(bold=True)
        ws.cell(tf, 3, sum(m.get("monto_guaranies", 0) for m in movimientos))
        ws.cell(tf, 4, sum(m.get("cantidad_fichas", 0) for m in movimientos))
        for col in range(1, 6):
            _estilo_total(ws.cell(tf, col))
        _fmt_moneda(ws.cell(tf, 3))

    _autofit(ws)


# ── Hoja 3: Ventas por Producto ────────────────────────────────────────────────

def _hoja_ventas_producto(wb, color: str,
                          fecha_inicio: str = None, fecha_fin: str = None):
    ws = wb.create_sheet("VENTAS POR PRODUCTO")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24

    headers = ["Producto", "Cantidad", "Total Fichas", "Total Guaraníes"]
    for col, h in enumerate(headers, 1):
        _estilo_header(ws.cell(1, col, h), color)

    if fecha_inicio and fecha_fin:
        ventas = ventas_por_producto_periodo(fecha_inicio, fecha_fin)
    else:
        ventas = ventas_por_producto()

    for fila, v in enumerate(ventas, start=2):
        fichas  = v.get("total_fichas", 0)
        guarani = fichas * VALOR_FICHA
        ws.cell(fila, 1, v.get("_id", ""))
        ws.cell(fila, 2, v.get("total_cantidad", 0))
        ws.cell(fila, 3, fichas)
        ws.cell(fila, 4, guarani)
        for col in range(1, 5):
            _estilo_celda(ws.cell(fila, col))
        _fmt_moneda(ws.cell(fila, 4))

    if ventas:
        tf         = len(ventas) + 2
        total_fic  = sum(v.get("total_fichas", 0) for v in ventas)
        ws.cell(tf, 1, "TOTALES")
        ws.cell(tf, 2, sum(v.get("total_cantidad", 0) for v in ventas))
        ws.cell(tf, 3, total_fic)
        ws.cell(tf, 4, total_fic * VALOR_FICHA)
        for col in range(1, 5):
            _estilo_total(ws.cell(tf, col))
        _fmt_moneda(ws.cell(tf, 4))

    _autofit(ws)


# ── Función principal ──────────────────────────────────────────────────────────

def generar_excel(tipo: str = "manual",
                  fecha_inicio: str = None,
                  fecha_fin: str = None) -> str:
    os.makedirs(REPORTS_DIR_VENTAS,     exist_ok=True)
    os.makedirs(REPORTS_DIR_PRODUCCION, exist_ok=True)

    nombre, titulo, color = _info_reporte(tipo, fecha_inicio, fecha_fin)

    wb = Workbook()
    wb.remove(wb.active)

    _hoja_resumen(wb, titulo, color, fecha_inicio, fecha_fin)
    _hoja_movimientos(wb, color, fecha_inicio, fecha_fin)
    _hoja_ventas_producto(wb, color, fecha_inicio, fecha_fin)

    ruta_ventas     = os.path.join(REPORTS_DIR_VENTAS,     nombre)
    ruta_produccion = os.path.join(REPORTS_DIR_PRODUCCION, nombre)

    wb.save(ruta_ventas)
    wb.save(ruta_produccion)

    logger.info(json.dumps({
        "event":   "reporte_generado",
        "tipo":    tipo,
        "nombre":  nombre,
        "periodo": f"{fecha_inicio}/{fecha_fin}" if fecha_inicio else "historico",
    }))
    return nombre
